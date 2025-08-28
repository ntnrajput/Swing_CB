import pandas as pd
from openpyxl import load_workbook

# ==== File paths ====
backtest_file = "backtest_df.xlsx"
features_file = "feature_backtest.csv"
output_file = "backtest_df_updated.xlsx"   # save as Excel (not CSV)

# ==== Read ====
df_backtest = pd.read_excel(backtest_file, sheet_name="Sheet1")
df_features = pd.read_csv(features_file)

# ---- Rename matching key columns ----
df_backtest.rename(columns={df_backtest.columns[1]: "date",
                            df_backtest.columns[2]: "symbol"}, inplace=True)
df_features.rename(columns={df_features.columns[7]: "date",
                            df_features.columns[6]: "symbol"}, inplace=True)

# ---- Convert dates ----
df_backtest["date"] = pd.to_datetime(df_backtest["date"], errors="coerce").dt.normalize()
df_features["date"] = pd.to_datetime(df_features["date"], errors="coerce", dayfirst=True).dt.normalize()

# ---- Get columns I to BV (8 to 85) ----
cols_to_copy = df_features.columns[8:85]

# ---- Merge ----
df_merged = pd.merge(
    df_backtest,
    df_features[["date", "symbol"] + list(cols_to_copy)],
    on=["date", "symbol"],
    how="left"
)

# ---- Place features from col N ----
N_index = 13
df_final = pd.concat([df_merged.iloc[:, :N_index], df_merged[cols_to_copy]], axis=1)

# ==== Save to Excel ====
df_final.to_excel(output_file, index=False)

# ==== Insert values in column L for BUY & delete SELL rows ====
wb = load_workbook(output_file)
ws = wb.active

# Find column indexes
col_action = 4   # D column
col_symbol = 3   # C column
col_pnl_pct = 12 # L column

max_row = ws.max_row
rows_to_delete = []

for row in range(2, max_row + 1):  # skip header
    action_value = ws.cell(row=row, column=col_action).value
    if action_value == "BUY":
        symbol = ws.cell(row=row, column=col_symbol).value
        found_value = None
        for search_row in range(row + 1, min(max_row, row + 1500) + 1):
            if ws.cell(search_row, column=col_symbol).value == symbol:
                found_value = ws.cell(search_row, column=col_pnl_pct).value
                break
        if found_value is not None:
            ws.cell(row=row, column=col_pnl_pct).value = found_value
    elif action_value == "SELL":
        rows_to_delete.append(row)

for r in reversed(rows_to_delete):
    ws.delete_rows(r, 1)

# ==== Insert 2 new rows after header ====
ws.insert_rows(2, amount=2)

# Recalculate max_row after deletion + insertion
max_row = ws.max_row
max_col = ws.max_column

# ==== Calculate min, max, avg for each feature column ====
# Convert worksheet to DataFrame for easy computation
data = pd.DataFrame(ws.values)
data.columns = data.iloc[0]   # first row is header
data = data.drop(0)           # remove header row

# Ensure numeric
data = data.apply(pd.to_numeric, errors="ignore")

# Column L = pnl_pct
pnl = pd.to_numeric(data.iloc[:, col_pnl_pct-1], errors="coerce")

# Two masks
mask_pos = pnl > 0
mask_neg = pnl < 0

for col in range(N_index, max_col + 1):  # from column N onward
    col_values = pd.to_numeric(data.iloc[:, col-1], errors="coerce")

    # Positive mask
    vals_pos = col_values[mask_pos].dropna()
    if not vals_pos.empty:
        min_v, max_v, avg_v = vals_pos.min(), vals_pos.max(), vals_pos.mean()
        ws.cell(row=2, column=col).value = f"{min_v:.2f}, {max_v:.2f}, {avg_v:.2f}"

    # Negative mask
    vals_neg = col_values[mask_neg].dropna()
    if not vals_neg.empty:
        min_v, max_v, avg_v = vals_neg.min(), vals_neg.max(), vals_neg.mean()
        ws.cell(row=3, column=col).value = f"{min_v:.2f}, {max_v:.2f}, {avg_v:.2f}"

wb.save(output_file)

print(f"✅ Updated Excel with BUY values, SELL rows deleted, and summary rows saved to {output_file}")
